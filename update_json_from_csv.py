#!/usr/bin/env python3

"""
Utility script to update JSON configuration file with normalization 
parameters from a Monte Carlo CSV or Excel results file.

This allows easy transfer of optimal normalization parameters from 
successful Monte Carlo runs to the JSON configuration used by 
recommend_model.py.
"""

import json
import csv
import click
import os
from typing import Dict, Any, Optional
import logging

from functions.logger_config import get_logger

logger = get_logger(__name__)


def read_csv_row(csv_file: str, row_number: int) -> Dict[str, Any]:
    """
    Read a specific row from the CSV file and extract normalization 
    parameters.
    
    Args:
        csv_file: Path to CSV file
        row_number: Row number to read (1-indexed, excluding header)
        
    Returns:
        Dictionary with central_values, std_values, lookback_days,
        performance_metrics, and metric_blending
    """
    with open(csv_file, 'r') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        
    if row_number < 1 or row_number > len(rows):
        raise ValueError(
            f"Row number {row_number} out of range. "
            f"CSV has {len(rows)} data rows."
        )
    
    row = rows[row_number - 1]
    
    # Extract central values
    central_values = {
        'annual_return': float(row['Central Annual Return']),
        'sharpe_ratio': float(row['Central Sharpe Ratio']),
        'sortino_ratio': float(row['Central Sortino Ratio']),
        'max_drawdown': float(row['Central Max Drawdown']),
        'avg_drawdown': float(row['Central Avg Drawdown'])
    }
    
    # Extract standard deviation values
    std_values = {
        'annual_return': float(row['Std Annual Return']),
        'sharpe_ratio': float(row['Std Sharpe Ratio']),
        'sortino_ratio': float(row['Std Sortino Ratio']),
        'max_drawdown': float(row['Std Max Drawdown']),
        'avg_drawdown': float(row['Std Avg Drawdown'])
    }
    
    # Extract lookback periods
    lookback_days = [
        int(row['Lookback Period 1']),
        int(row['Lookback Period 2']),
        int(row['Lookback Period 3'])
    ]
    
    # Extract performance metric weights
    performance_metrics = {
        'sharpe_ratio_weight': float(row['Sharpe Ratio Weight']),
        'sortino_ratio_weight': float(row['Sortino Ratio Weight']),
        'max_drawdown_weight': float(row['Max Drawdown Weight']),
        'avg_drawdown_weight': float(row['Avg Drawdown Weight']),
        'annualized_return_weight': float(row['Annualized Return Weight'])
    }
    
    # Extract metric blending parameters
    metric_blending = {
        'enabled': row['Metric Blending Enabled'].lower() == 'true',
        'full_period_weight': float(row['Full Period Weight']),
        'focus_period_weight': float(row['Focus Period Weight'])
    }
    
    return {
        'central_values': central_values,
        'std_values': std_values,
        'lookback_days': lookback_days,
        'performance_metrics': performance_metrics,
        'metric_blending': metric_blending,
        'row_data': row
    }


def find_row_by_datetime(csv_file: str, date: str, time: str) -> Dict[str, Any]:
    """
    Find a row in the CSV file by date and time.
    
    Args:
        csv_file: Path to CSV file
        date: Date string to search for (e.g., '2025-09-14')
        time: Time string to search for (e.g., '22:36:59')
        
    Returns:
        Dictionary with central_values, std_values, lookback_days,
        performance_metrics, metric_blending, and row_data
        
    Raises:
        ValueError: If no matching row is found
    """
    with open(csv_file, 'r') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    
    for idx, row in enumerate(rows, start=1):
        if row['Date'] == date and row['Time'] == time:
            logger.info(f"Found matching row at index {idx}")
            
            # Extract central values
            central_values = {
                'annual_return': float(row['Central Annual Return']),
                'sharpe_ratio': float(row['Central Sharpe Ratio']),
                'sortino_ratio': float(row['Central Sortino Ratio']),
                'max_drawdown': float(row['Central Max Drawdown']),
                'avg_drawdown': float(row['Central Avg Drawdown'])
            }
            
            # Extract standard deviation values
            std_values = {
                'annual_return': float(row['Std Annual Return']),
                'sharpe_ratio': float(row['Std Sharpe Ratio']),
                'sortino_ratio': float(row['Std Sortino Ratio']),
                'max_drawdown': float(row['Std Max Drawdown']),
                'avg_drawdown': float(row['Std Avg Drawdown'])
            }
            
            # Extract lookback periods
            lookback_days = [
                int(row['Lookback Period 1']),
                int(row['Lookback Period 2']),
                int(row['Lookback Period 3'])
            ]
            
            # Extract performance metric weights
            performance_metrics = {
                'sharpe_ratio_weight': float(row['Sharpe Ratio Weight']),
                'sortino_ratio_weight': float(row['Sortino Ratio Weight']),
                'max_drawdown_weight': float(row['Max Drawdown Weight']),
                'avg_drawdown_weight': float(row['Avg Drawdown Weight']),
                'annualized_return_weight': float(row['Annualized Return Weight'])
            }
            
            # Extract metric blending parameters
            metric_blending = {
                'enabled': row['Metric Blending Enabled'].lower() == 'true',
                'full_period_weight': float(row['Full Period Weight']),
                'focus_period_weight': float(row['Focus Period Weight'])
            }
            
            return {
                'central_values': central_values,
                'std_values': std_values,
                'lookback_days': lookback_days,
                'performance_metrics': performance_metrics,
                'metric_blending': metric_blending,
                'row_data': row,
                'row_number': idx
            }
    
    raise ValueError(
        f"No row found with Date='{date}' and Time='{time}'. "
        f"CSV has {len(rows)} total rows."
    )


def read_excel_row(excel_file: str, row_number: int) -> Dict[str, Any]:
    """
    Read a specific row from an Excel file and extract normalization 
    parameters.
    
    Args:
        excel_file: Path to Excel file
        row_number: Row number to read (1-indexed, excluding header)
        
    Returns:
        Dictionary with central_values, std_values, lookback_days,
        performance_metrics, and metric_blending
    """
    try:
        import openpyxl
        from datetime import datetime
    except ImportError:
        raise ImportError(
            "openpyxl is required to read Excel files. "
            "Install it with: uv add openpyxl"
        )
    
    def safe_float_convert(value):
        """Convert value to float, handling datetime objects."""
        if isinstance(value, datetime):
            # If it's a datetime, Excel likely misformatted a decimal as a time
            # Extract the numeric value from the datetime
            # Excel stores dates as days since 1900-01-01
            # So datetime(1900, 1, 1, 9, 36, 0) = day 1 + 0.4 days = 1.4
            logger.warning(
                f"Found datetime value where number expected: {value}. "
                f"This indicates incorrect Excel cell formatting. "
                f"Converting to numeric value."
            )
            # Convert to Excel serial number (days since 1900-01-01)
            # Then extract just the decimal part if it's close to day 1
            from datetime import datetime as dt
            excel_epoch = dt(1899, 12, 30)  # Excel's epoch
            delta = value - excel_epoch
            numeric_value = delta.days + (delta.seconds / 86400.0)
            
            # If the value is small (< 100), it's likely a misformatted decimal
            if numeric_value < 100:
                return numeric_value
            else:
                # For larger values, just return the fractional day part
                return (delta.seconds / 86400.0) + delta.days
        elif value is None:
            raise ValueError("Cannot convert None to float")
        else:
            return float(value)
    
    workbook = openpyxl.load_workbook(excel_file, read_only=True)
    sheet = workbook.active
    
    # Header is at row 13, data starts at row 14
    # Read up to column 42 (indices 0-41) to include all data columns
    header_row = list(sheet.iter_rows(min_row=13, max_row=13, 
                                       max_col=42, values_only=True))[0]
    
    # Get all data rows starting from row 14
    rows = list(sheet.iter_rows(min_row=14, max_col=42, values_only=True))
    
    if row_number < 1 or row_number > len(rows):
        raise ValueError(
            f"Row number {row_number} out of range. "
            f"Excel file has {len(rows)} data rows."
        )
    
    row_data = rows[row_number - 1]
    
    # Create dictionary from row data
    row = {header_row[i]: row_data[i] for i in range(len(header_row)) 
           if header_row[i] is not None}
    
    # Extract central values with safe conversion
    central_values = {
        'annual_return': safe_float_convert(row['Central Annual Return']),
        'sharpe_ratio': safe_float_convert(row['Central Sharpe Ratio']),
        'sortino_ratio': safe_float_convert(row['Central Sortino Ratio']),
        'max_drawdown': safe_float_convert(row['Central Max Drawdown']),
        'avg_drawdown': safe_float_convert(row['Central Avg Drawdown'])
    }
    
    # Extract standard deviation values with safe conversion
    std_values = {
        'annual_return': safe_float_convert(row['Std Annual Return']),
        'sharpe_ratio': safe_float_convert(row['Std Sharpe Ratio']),
        'sortino_ratio': safe_float_convert(row['Std Sortino Ratio']),
        'max_drawdown': safe_float_convert(row['Std Max Drawdown']),
        'avg_drawdown': safe_float_convert(row['Std Avg Drawdown'])
    }
    
    # Extract lookback periods
    lookback_days = [
        int(row['Lookback Period 1']),
        int(row['Lookback Period 2']),
        int(row['Lookback Period 3'])
    ]
    
    # Extract performance metric weights with safe conversion
    performance_metrics = {
        'sharpe_ratio_weight': safe_float_convert(row['Sharpe Ratio Weight']),
        'sortino_ratio_weight': safe_float_convert(row['Sortino Ratio Weight']),
        'max_drawdown_weight': safe_float_convert(row['Max Drawdown Weight']),
        'avg_drawdown_weight': safe_float_convert(row['Avg Drawdown Weight']),
        'annualized_return_weight': safe_float_convert(row['Annualized Return Weight'])
    }
    
    # Extract metric blending parameters with safe conversion
    metric_blending = {
        'enabled': str(row['Metric Blending Enabled']).lower() == 'true',
        'full_period_weight': safe_float_convert(row['Full Period Weight']),
        'focus_period_weight': safe_float_convert(row['Focus Period Weight'])
    }
    
    workbook.close()
    
    return {
        'central_values': central_values,
        'std_values': std_values,
        'lookback_days': lookback_days,
        'performance_metrics': performance_metrics,
        'metric_blending': metric_blending,
        'row_data': row
    }


def find_excel_row_by_datetime(excel_file: str, date: str, 
                               time: str) -> Dict[str, Any]:
    """
    Find a row in an Excel file by date and time.
    
    Args:
        excel_file: Path to Excel file
        date: Date string to search for (e.g., '2025-09-14')
        time: Time string to search for (e.g., '22:36:59')
        
    Returns:
        Dictionary with central_values, std_values, lookback_days,
        performance_metrics, metric_blending, row_data, and row_number
        
    Raises:
        ValueError: If no matching row is found
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required to read Excel files. "
            "Install it with: uv add openpyxl"
        )
    
    workbook = openpyxl.load_workbook(excel_file, read_only=True)
    sheet = workbook.active
    
    # Header is at row 13, only use columns A through AE (indices 0-30)
    header_row = list(sheet.iter_rows(min_row=13, max_row=13, 
                                       max_col=42, values_only=True))[0]
    
    # Find Date and Time column indices
    date_col = None
    time_col = None
    for idx, col_name in enumerate(header_row):
        if col_name == 'Date':
            date_col = idx
        elif col_name == 'Time':
            time_col = idx
    
    if date_col is None or time_col is None:
        workbook.close()
        raise ValueError("Could not find 'Date' or 'Time' columns in Excel file")
    
    # Search for matching row (data starts at row 14)
    for row_idx, row_data in enumerate(
        sheet.iter_rows(min_row=14, max_col=42, values_only=True), start=1
    ):
        row_date = str(row_data[date_col]) if row_data[date_col] else ""
        row_time = str(row_data[time_col]) if row_data[time_col] else ""
        
        if row_date == date and row_time == time:
            logger.info(f"Found matching row at index {row_idx}")
            
            # Create dictionary from row data
            row = {header_row[i]: row_data[i] for i in range(len(header_row)) 
                   if header_row[i] is not None}
            
            # Extract central values
            central_values = {
                'annual_return': float(row['Central Annual Return']),
                'sharpe_ratio': float(row['Central Sharpe Ratio']),
                'sortino_ratio': float(row['Central Sortino Ratio']),
                'max_drawdown': float(row['Central Max Drawdown']),
                'avg_drawdown': float(row['Central Avg Drawdown'])
            }
            
            # Extract standard deviation values
            std_values = {
                'annual_return': float(row['Std Annual Return']),
                'sharpe_ratio': float(row['Std Sharpe Ratio']),
                'sortino_ratio': float(row['Std Sortino Ratio']),
                'max_drawdown': float(row['Std Max Drawdown']),
                'avg_drawdown': float(row['Std Avg Drawdown'])
            }
            
            # Extract lookback periods
            lookback_days = [
                int(row['Lookback Period 1']),
                int(row['Lookback Period 2']),
                int(row['Lookback Period 3'])
            ]
            
            # Extract performance metric weights
            performance_metrics = {
                'sharpe_ratio_weight': float(row['Sharpe Ratio Weight']),
                'sortino_ratio_weight': float(row['Sortino Ratio Weight']),
                'max_drawdown_weight': float(row['Max Drawdown Weight']),
                'avg_drawdown_weight': float(row['Avg Drawdown Weight']),
                'annualized_return_weight': float(row['Annualized Return Weight'])
            }
            
            # Extract metric blending parameters
            metric_blending = {
                'enabled': str(row['Metric Blending Enabled']).lower() == 'true',
                'full_period_weight': float(row['Full Period Weight']),
                'focus_period_weight': float(row['Focus Period Weight'])
            }
            
            workbook.close()
            
            return {
                'central_values': central_values,
                'std_values': std_values,
                'lookback_days': lookback_days,
                'performance_metrics': performance_metrics,
                'metric_blending': metric_blending,
                'row_data': row,
                'row_number': row_idx
            }
    
    workbook.close()
    
    raise ValueError(
        f"No row found with Date='{date}' and Time='{time}' in Excel file"
    )


def update_json_config(
    json_file: str, 
    central_values: Dict[str, float],
    std_values: Dict[str, float],
    lookback_days: list,
    performance_metrics: Dict[str, float],
    metric_blending: Dict[str, Any],
    backup: bool = True
) -> None:
    """
    Update JSON configuration file with new parameters from CSV/Excel.
    
    Args:
        json_file: Path to JSON configuration file
        central_values: Dictionary of central values
        std_values: Dictionary of standard deviation values
        lookback_days: List of lookback periods
        performance_metrics: Dictionary of performance metric weights
        metric_blending: Dictionary of metric blending parameters
        backup: Whether to create backup before modifying
    """
    if backup:
        backup_file = f"{json_file}.backup"
        with open(json_file, 'r') as f:
            backup_content = f.read()
        with open(backup_file, 'w') as f:
            f.write(backup_content)
        logger.info(f"Created backup at {backup_file}")
    
    with open(json_file, 'r') as f:
        config = json.load(f)
    
    # Update model_selection section
    if 'model_selection' not in config:
        config['model_selection'] = {}
    
    # Update normalization parameters
    if 'normalization' not in config['model_selection']:
        config['model_selection']['normalization'] = {}
    
    config['model_selection']['normalization']['central_values'] = \
        central_values
    config['model_selection']['normalization']['std_values'] = std_values
    config['model_selection']['normalization']['lookback_days'] = lookback_days
    
    # Update performance metrics weights
    config['model_selection']['performance_metrics'] = performance_metrics
    
    # Update metric blending parameters
    if 'metric_blending' not in config:
        config['metric_blending'] = {}
    
    config['metric_blending']['enabled'] = metric_blending['enabled']
    config['metric_blending']['full_period_weight'] = \
        metric_blending['full_period_weight']
    config['metric_blending']['focus_period_weight'] = \
        metric_blending['focus_period_weight']
    
    #############################################################################
    # Update recommendation_mode.default_lookbacks for recommend_model.py
    #############################################################################
    if 'recommendation_mode' not in config:
        config['recommendation_mode'] = {}
    
    config['recommendation_mode']['default_lookbacks'] = lookback_days
    
    with open(json_file, 'w') as f:
        json.dump(config, f, indent=4)
    
    logger.info(f"Updated {json_file} with all parameters from CSV/Excel row")
    logger.info(f"  - Normalization values (central and std)")
    logger.info(f"  - Lookback days: {lookback_days}")
    logger.info(f"  - Performance metric weights")
    logger.info(f"  - Metric blending parameters")
    logger.info(f"  - Recommendation mode default lookbacks")


@click.command()
@click.option(
    '--csv', 'csv_file',
    required=False,
    type=click.Path(exists=True),
    help='Path to Monte Carlo results CSV file'
)
@click.option(
    '--xlsx', 'xlsx_file',
    required=False,
    type=click.Path(exists=True),
    help='Path to Monte Carlo results Excel file'
)
@click.option(
    '--row', 'row_number',
    required=False,
    type=int,
    help='Excel row number as shown in Excel (e.g., 419 for row 419)'
)
@click.option(
    '--excel-row', 'excel_row_number',
    required=False,
    type=int,
    help='(Deprecated - use --row instead) Actual Excel row number'
)
@click.option(
    '--date', 'date',
    required=False,
    type=str,
    help='Date to search for in the file (e.g., "2025-09-14")'
)
@click.option(
    '--time', 'time',
    required=False,
    type=str,
    help='Time to search for in the file (e.g., "22:36:59")'
)
@click.option(
    '--json', 'json_file',
    required=True,
    type=click.Path(exists=True),
    help='Path to JSON configuration file to update'
)
@click.option(
    '--no-backup',
    is_flag=True,
    help='Skip creating backup of JSON file before updating'
)
@click.option(
    '--show-only',
    is_flag=True,
    help='Show parameters without updating JSON file'
)
def main(
    csv_file: Optional[str],
    xlsx_file: Optional[str],
    row_number: Optional[int],
    excel_row_number: Optional[int],
    date: Optional[str],
    time: Optional[str],
    json_file: str,
    no_backup: bool,
    show_only: bool
) -> None:
    """
    Update JSON configuration with normalization parameters from CSV or Excel row.
    
    IMPORTANT: Row numbering differs between CSV and Excel files:
    - CSV files: --row is 1-indexed from first data row (after header)
    - Excel files: --row is also 1-indexed from first data row (row 14 in Excel)
                   --excel-row is the actual Excel row number (e.g., 419)
    
    Example usage:
    
        # From CSV file by row number
        uv run python update_json_from_csv.py \\
            --csv abacus_best_performers.csv \\
            --row 42 \\
            --json pytaaa_model_switching_params.json
        
        # From Excel file by data row number (row 42 = Excel row 55)
        uv run python update_json_from_csv.py \\
            --xlsx abacus_best_performers_3.xlsx \\
            --row 42 \\
            --json pytaaa_model_switching_params.json
        
        # From Excel file by actual Excel row number
        uv run python update_json_from_csv.py \\
            --xlsx abacus_best_performers_3.xlsx \\
            --excel-row 419 \\
            --json pytaaa_model_switching_params.json
        
        # From Excel file by date and time
        uv run python update_json_from_csv.py \\
            --xlsx abacus_best_performers_3.xlsx \\
            --date "2025-09-14" \\
            --time "22:36:59" \\
            --json pytaaa_model_switching_params.json
    """
    try:
        # Validate input arguments
        if not csv_file and not xlsx_file:
            raise ValueError("Either --csv or --xlsx must be provided")
        
        if csv_file and xlsx_file:
            raise ValueError("Cannot specify both --csv and --xlsx")
        
        if excel_row_number and csv_file:
            raise ValueError("--excel-row can only be used with Excel files (--xlsx)")
        
        if row_number and excel_row_number:
            raise ValueError("Cannot specify both --row and --excel-row")
        
        # Determine file type and read data
        if csv_file:
            file_type = "CSV"
            source_file = csv_file
            print(f"Reading CSV file: {csv_file}")
            
            if row_number:
                print(f"Extracting parameters from data row: {row_number}")
                params = read_csv_row(csv_file, row_number)
            elif date and time:
                print(f"Searching for row with Date='{date}' and Time='{time}'")
                params = find_row_by_datetime(csv_file, date, time)
                row_number = params['row_number']
            else:
                raise ValueError("Either --row or both --date and --time must be provided")
        
        else:  # xlsx_file
            file_type = "Excel"
            source_file = xlsx_file
            print(f"Reading Excel file: {xlsx_file}")
            
            # Handle deprecated --excel-row option
            if excel_row_number:
                print("Warning: --excel-row is deprecated, use --row instead")
                row_number = excel_row_number
            
            if row_number:
                # For Excel files, convert actual Excel row number to data row number
                # Header is at row 13, data starts at row 14
                if row_number < 14:
                    raise ValueError(
                        f"Excel row {row_number} is before data starts (row 14)"
                    )
                data_row = row_number - 13
                print(f"Extracting parameters from Excel row {row_number}")
                params = read_excel_row(xlsx_file, data_row)
            elif date and time:
                print(f"Searching for row with Date='{date}' and Time='{time}'")
                params = find_excel_row_by_datetime(xlsx_file, date, time)
                data_row = params['row_number']
                row_number = data_row + 13
                print(f"Found at Excel row {row_number}")
            else:
                raise ValueError(
                    "Either --row or both --date and --time must be provided"
                )
        
        print("\n" + "="*60)
        print("EXTRACTED PARAMETERS")
        print("="*60)
        print(f"Source: {file_type} file - Row {row_number}")
        
        print("\nCentral Values:")
        for key, value in params['central_values'].items():
            print(f"  {key}: {value}")
        
        print("\nStandard Deviation Values:")
        for key, value in params['std_values'].items():
            print(f"  {key}: {value}")
        
        print("\nSource Row Information:")
        print(f"  Date: {params['row_data']['Date']}")
        print(f"  Time: {params['row_data']['Time']}")
        
        # Format Final Value with $ and commas
        final_value_raw = params['row_data']['Final Value']
        if isinstance(final_value_raw, str):
            # Remove existing $ and commas if present
            final_value_clean = final_value_raw.replace('$', '').replace(',', '')
            final_value = float(final_value_clean)
        else:
            final_value = float(final_value_raw)
        
        print(f"  Final Value: ${final_value:,.0f}")
        print(f"  Annual Return: {params['row_data']['Annual Return']}")
        print(f"  Sharpe Ratio: {params['row_data']['Sharpe Ratio']}")
        print(f"  Sortino Ratio: {params['row_data']['Sortino Ratio']}")
        print(f"  Max Drawdown: {params['row_data']['Max Drawdown']}")
        
        print("\nLookback Periods:")
        print(f"  {params['lookback_days']}")
        
        print("\nPerformance Metric Weights:")
        for key, value in params['performance_metrics'].items():
            print(f"  {key}: {value}")
        
        print("\nMetric Blending:")
        print(f"  enabled: {params['metric_blending']['enabled']}")
        print(f"  full_period_weight: {params['metric_blending']['full_period_weight']}")
        print(f"  focus_period_weight: {params['metric_blending']['focus_period_weight']}")
        
        if show_only:
            print("\n" + "="*60)
            print("--show-only flag set, not updating JSON file")
            print("="*60)
            return
        
        print("\n" + "="*60)
        print(f"Updating JSON file: {json_file}")
        print("="*60)
        
        update_json_config(
            json_file,
            params['central_values'],
            params['std_values'],
            params['lookback_days'],
            params['performance_metrics'],
            params['metric_blending'],
            backup=not no_backup
        )
        
        print(f"\n✓ Successfully updated {json_file}")
        if not no_backup:
            print(f"  Backup saved as {json_file}.backup")
        
    except Exception as e:
        logger.error(f"Error updating JSON from file: {str(e)}", exc_info=True)
        print(f"\n✗ Error: {str(e)}")
        raise click.Abort()


if __name__ == "__main__":
    main()
