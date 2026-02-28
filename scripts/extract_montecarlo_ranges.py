#!/usr/bin/env python3
"""Extract triangular-range constants from Monte Carlo optimisation xlsx files.

This is a developer utility script.  Run it once after completing an
xlsx-based parameter optimisation, then paste the printed ``_RANGES``
dict into ``functions/backtesting/parameter_exploration.py``.

The xlsx files are never read at runtime \u2014 ranges are baked into the
source code as constants.

Usage
-----
::

    uv run python scripts/extract_montecarlo_ranges.py

Output
------
A ready-to-paste Python ``_RANGES`` dict literal, printed to stdout.

Xlsx structure expected
-----------------------
- Column B : statistic label (P10, mean, median, P90)
- Row 2    : P10 values
- Row 3    : mean values
- Row 4    : median values
- Row 5    : P90 values
- Row 9    : column headers matching parameter names in the trial dict
- Columns C onward: one column per parameter
"""

import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(
        "ERROR: openpyxl is required.  Install with:  uv add openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)

#############################################################################
# Scenario \u2192 xlsx directory mapping
#############################################################################

_XLSX_FILENAME = "pytaaa_backtest_montecarlo.xlsx"

_SCENARIO_DIRS: dict[str, str] = {
    "naz100_pine": "/Users/donaldpg/pyTAAA_data/naz100_pine/pytaaa_backtest/",
    "naz100_hma":  "/Users/donaldpg/pyTAAA_data/naz100_hma/pytaaa_backtest/",
    "naz100_pi":   "/Users/donaldpg/pyTAAA_data/naz100_pi/pytaaa_backtest/",
    "sp500_pine":  "/Users/donaldpg/pyTAAA_data/sp500_pine/pytaaa_backtest/",
    "sp500_hma":   "/Users/donaldpg/pyTAAA_data/sp500_hma/pytaaa_backtest/",
}

# Row indices (1-based, as openpyxl uses)
_ROW_P10    = 2
_ROW_MEAN   = 3
_ROW_MEDIAN = 4
_ROW_P90    = 5
_ROW_HEADER = 9

# Parameters that use triangular sampling (randint params excluded).
_TRIANGULAR_PARAMS = (
    "LongPeriod",
    "stddevThreshold",
    "MA1",
    "MA2",
    "MA2offset",
    "sma2factor",
    "rankThresholdPct",
    "riskDownside_min",
    "riskDownside_max",
    "sma_filt_val",
    "lowPct",
    "hiPct",
)

# Fallback (low, mid, high) used when the xlsx is missing or a column
# is absent.  These match the initial values in parameter_exploration.py.
_FALLBACK: dict[str, tuple] = {
    "LongPeriod":       (150,   300,    600),
    "stddevThreshold":  (2.0,   8.0,    20.0),
    "MA1":              (75,    151,    300),
    "MA2":              (7,     22,     45),
    "MA2offset":        (2,     7,      12),
    "sma2factor":       (1.65,  2.5,    2.75),
    "rankThresholdPct": (0.14,  0.20,   0.26),
    "riskDownside_min": (0.50,  0.70,   0.90),
    "riskDownside_max": (8.0,   10.0,   13.0),
    "sma_filt_val":     (0.010, 0.015,  0.0225),
    "lowPct":           (10.0,  20.0,   30.0),
    "hiPct":            (70.0,  80.0,   90.0),
}


def _extract_ranges(xlsx_path: str, scenario: str) -> dict[str, tuple]:
    """Extract (low, mid, high) triples from a single xlsx file.

    Args:
        xlsx_path: Absolute path to the xlsx file.
        scenario: Scenario name (used in warning messages only).

    Returns:
        Dict mapping parameter names to (P10, avg(mean,median), P90).
        Missing columns fall back to ``_FALLBACK`` values.
    """
    if not os.path.isfile(xlsx_path):
        print(
            f"  WARNING: {xlsx_path} not found \u2014 using fallback values.",
            file=sys.stderr,
        )
        return dict(_FALLBACK)

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        print(
            f"  WARNING: Could not open {xlsx_path}: {exc} "
            f"\u2014 using fallback values.",
            file=sys.stderr,
        )
        return dict(_FALLBACK)

    # Read header row to map column index \u2192 parameter name.
    header_row = list(ws.iter_rows(
        min_row=_ROW_HEADER, max_row=_ROW_HEADER, values_only=True
    ))[0]
    # Column B (index 1) is the stat label; params start at column C (index 2).
    col_to_param: dict[int, str] = {}
    for col_idx, cell_val in enumerate(header_row):
        if col_idx < 2:
            continue
        if cell_val is not None:
            col_to_param[col_idx] = str(cell_val).strip()

    def _row_values(row_num: int) -> dict[str, float]:
        """Read a data row and return {param_name: value}."""
        row = list(ws.iter_rows(
            min_row=row_num, max_row=row_num, values_only=True
        ))[0]
        result = {}
        for col_idx, param_name in col_to_param.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    result[param_name] = float(row[col_idx])
                except (TypeError, ValueError):
                    pass
        return result

    p10_vals    = _row_values(_ROW_P10)
    mean_vals   = _row_values(_ROW_MEAN)
    median_vals = _row_values(_ROW_MEDIAN)
    p90_vals    = _row_values(_ROW_P90)

    wb.close()

    ranges: dict[str, tuple] = {}
    for param in _TRIANGULAR_PARAMS:
        if param not in p10_vals or param not in p90_vals:
            print(
                f"  WARNING [{scenario}]: column '{param}' missing "
                f"from xlsx \u2014 using fallback.",
                file=sys.stderr,
            )
            ranges[param] = _FALLBACK[param]
            continue
        low = p10_vals[param]
        high = p90_vals[param]
        mean = mean_vals.get(param, (low + high) / 2.0)
        median = median_vals.get(param, (low + high) / 2.0)
        mid = (mean + median) / 2.0
        # Clamp mid to [low, high] in case of xlsx data anomalies.
        mid = max(low, min(mid, high))
        ranges[param] = (low, mid, high)

    return ranges


def _format_ranges_dict(all_ranges: dict[str, dict[str, tuple]]) -> str:
    """Format all scenario ranges as a ready-to-paste Python dict literal.

    Args:
        all_ranges: {scenario: {param: (low, mid, high)}}

    Returns:
        Formatted Python source string.
    """
    lines = ["_RANGES: dict[str, dict[str, tuple]] = {"]
    for scenario, ranges in all_ranges.items():
        lines.append(f'    "{scenario}": {{')
        # Align columns for readability.
        max_key_len = max(len(k) for k in ranges)
        for param, (low, mid, high) in ranges.items():
            pad = " " * (max_key_len - len(param))
            # Format numbers: use repr to preserve float precision cleanly.
            lines.append(
                f'        "{param}":{pad} ({low!r:>8}, {mid!r:>8}, {high!r}),',
            )
        lines.append("    },")
    lines.append("}")
    return "\n".join(lines)


def main() -> None:
    """Extract ranges from all scenario xlsx files and print the result."""
    print("Extracting ranges from xlsx files...\n", file=sys.stderr)

    all_ranges: dict[str, dict[str, tuple]] = {}
    for scenario, directory in _SCENARIO_DIRS.items():
        xlsx_path = os.path.join(directory, _XLSX_FILENAME)
        print(f"  [{scenario}] {xlsx_path}", file=sys.stderr)
        all_ranges[scenario] = _extract_ranges(xlsx_path, scenario)

    print("\n# ---- PASTE THE FOLLOWING INTO parameter_exploration.py ----\n")
    print(_format_ranges_dict(all_ranges))
    print("\n# ---- END PASTE ----")


if __name__ == "__main__":
    main()
